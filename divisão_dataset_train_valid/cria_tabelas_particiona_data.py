import os
import random
import shutil
from openpyxl import Workbook, load_workbook# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
from globox import BoundingBox, AnnotationSet, BoxFormat
import numpy as np
############################ labels ##############################


label_to_id={'0':'limpeza_lona_carregadora', 
             '1':'descarte_biscoitos', 
             '2':'limpeza_gaveta',
             '3':'retirada_bobina',
             '4':'reposicao_bobina',
             '5':'filme_partindo_desalinhando',
             '6':'potenciometro_regulagem',
             '7':'organizacao_embalagens_unitarias',
             '8':'analize_mezanino',
             '9':'troca_teflon',
             '10':'umedecimento_correias_mageadoras',
             '11':'troca_correias_mageadoras',
             '12':'alinhamento_correias_mageadoras',
             '13':'ajuste_posicao_bags',
             '14':'interacao_mesa',
             '15':'ihm_bosch_vertical',
             '16':'ihm_mezanino',
             '17':'ihm_pak_mak',
             '18':'limpeza_piso',
             '19':'limpeza_caixa',
             '20':'desenrosco_biscoito',
             '21':'idle',
             '22':'encaixotamento',
             '23':'despejo_bandeja',
             '24':'limpeza_bag_vazio',
             '25':'ajuste_bobina',
             '26':'retirada_embalagens_unitarias',
             '27':'limpeza_pistola',
             '28':'limpeza_alcool',
             '29':'limpeza_mordente'
             }

##################################################################

#excel
# Cria Novo workbook
wb = Workbook()# Seleciona a active Sheet
ws1 = wb.active
# Rename it
ws1.title = 'my test'# Escreve alguns dados

############

nome_label='29/train'
directorio_pasta_label='/media/ernesto/ANTONIETTA/data-vita-saci-rgb/'
os.chdir(directorio_pasta_label)
HOME = os.getcwd()
print("HOME:", HOME)
dir_path = directorio_pasta_label+nome_label+'/'


col=2
colt=7
linha=2
linha3=2
linhav=5
ws1['A1']='LABEL'
ws1['B1']='VIDEO'
ws1['C1']='FRAMES'
ws1['F3']='TOTAL'
ws1['F4']='TREINO'
ws1['F5']='VALIDAÇÃO'

ws1['A2']=nome_label
DIR_T=[]
frames_total = []
totalG=[]

for (dirpath, dirnames, filenames) in os.walk(dir_path): # obtendo caminho atual, diretorios e ficheiros respetivamente
    #print(dirnames)
    
    for dir  in dirnames:
            print(dir)
            DIR_T.append(int(dir))

            nome_video=dir
            labels_folder=HOME+'/'+ nome_label+'/' +nome_video +'/'
            images_folder=HOME+'/'+ nome_label+'/' +nome_video +'/'
            print(labels_folder)
            yolo = AnnotationSet.from_yolo_v5(
                folder=labels_folder,
                
                image_folder=images_folder,

            )
            gts=yolo
            table=gts.show_stats(label_fora=label_to_id)

            from rich.console import Console
            console = Console(record=True)
            console.print(table, justify="center")            
            console.save_svg(nome_video+".svg", title=nome_video)

            import os

            letter = get_column_letter(col)
            #print(letter)
            ws1[letter+str(linha)] = dir# Cria nova sheet
            linha+=1           
            dir_pat = dir_path +dir
            #print(dir_pat)
            i=0
            #totalL=[]
            for file in os.listdir(dir_pat): # percorrer ficheiros em cada diretorio (dirpath)
                #print(file)
                i=i+1

            totalG.append(int(i/2))
        
for it in totalG:
    #print(it)
    letter = get_column_letter(col+1)
    ws1[letter+str(linha3)] = it# Cria nova sheet
    linha3=linha3+1

totalFrames = sum(totalG)
totalFrames_train= totalFrames*0.7

ws1['G3']=totalFrames
                 

print('Pastas dos video')
print(DIR_T)
numero=len(DIR_T) #numero de pastas po label
numero_pastas_valid=int(numero/3)
print('total de frames por video')
print(totalG)

print('numero total de frames da label')
print(sum(totalG))
print('totalG_35')
totalG_35=int(sum(totalG)*0.31)
print(totalG_35)
print('totalG_25')
totalG_25=int(sum(totalG)*0.25)
print(totalG_25)
#resultado_esco=[(totalG.index(x), x.index(19)) for x in totalG if 19 in x]

##################################################################
###### COMENTAR SE NÃO FOR TIVER UMA PASTA COM MUITOS FRAMES #####
# print('elemento muito grande a seer retirado')
# tirando_maior=totalG.index(117)
# tirando_maior_pasta=DIR_T.index(733)

# print(tirando_maior)
# del totalG[tirando_maior]
# print('new TotalG')
# print(totalG)
# del DIR_T[tirando_maior_pasta]
# print('new DIRT')
# print(DIR_T)

##################################################################
#print(resultado_esco)

VALID = random.sample(DIR_T,numero_pastas_valid)

i_g=[]
for val in VALID:
        for dirtt in  range(len(DIR_T)):
              #print(dirtt)
              if DIR_T[dirtt]==val:
                    i_g.append(dirtt)
                    
print("VALID")
print(VALID)
print('vetor com as posiçoes dois Numeros sorteados')
print(i_g) #vetor com as posiçoes dois Numeros sorteados

totalf_v=[]
for gi in i_g:

      #print(gi)
      totalf_v.append(totalG[gi])

print('Vetor com quantidade de frames sorteados')
print(totalf_v)
print('quantidade de frames sorteados')
print(sum(totalf_v))
ws1['G5']=sum(totalf_v)
ws1['G4']=totalFrames-sum(totalf_v) #total frames train
colv=8    
if sum(totalf_v)>=int(totalG_25) and sum(totalf_v) <= int(totalG_35):
        print("FOI SUCESSO")        
        print("FOI SUCESSO")  
        for val in VALID:
                letter = get_column_letter(colv)
                ws1[letter+str(linhav)] = val# Cria nova sheet
                linhav=linhav
                colv=colv+1
                ws1[letter+str(linhav-3)]='#VIDEO'
else:
      VALID = random.sample(DIR_T,numero_pastas_valid)
      
wb.save(nome_label+'.'+'xlsx')